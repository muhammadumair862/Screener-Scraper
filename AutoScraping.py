from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox
import os

# Selenium Webdriver settings
options = webdriver.ChromeOptions()
s = Service('chromedriver/chromedriver96.exe')
# Get the base path of the current script
base_path = os.path.dirname(os.path.abspath(__file__))
# file_path = os.path.join(base_path,'Book1.xlsx')
file_path = 'Book1.xlsx'



# Function to parse page using BeautifulSoup
def get_soup(response):
    return BeautifulSoup(response, 'html.parser')


# Function to login page using Selenium Webdriver
def login(driver, url='https://www.screener.in/login/', 
          username='vijata5689@vaband.com', password='qwerty@123'):
    # Navigate to the login page
    driver.get(url)

    # Find the email and password input fields
    email_input = driver.find_element(By.ID, 'id_username')
    password_input = driver.find_element(By.ID, 'id_password')

    # Enter your email and password
    email_input.send_keys(username)
    password_input.send_keys(password)

    # Submit the login form
    password_input.send_keys(Keys.RETURN)


# Get values from the page
def extract_values(page, values_dict):
    for i in page.find_all('li'):
        parameter_name = i.find('span', attrs={'class':'name'}).text.strip()
        try:
            if len(i.find_all('span', attrs={'class':'number'}))==2:
                parameter_val = []
                for k in i.find_all('span', attrs={'class':'number'}):
                    parameter_val.append(float(k.text.strip().replace(',', '')))
            else:        
                parameter_val = float(i.find('span', attrs={'class':'number'}).text.strip().replace(',', ''))
        except:
            parameter_val = i.find('span', attrs={'class':'number'}).text.strip().replace(',', '')
        # Check if the key already exists in the dictionary
        if parameter_name in values_dict:
            try:
                values_dict[parameter_name].append(parameter_val)
            except:
                values_dict[parameter_name].append('')
        else:
            try:
                values_dict[parameter_name] = [parameter_val]
            except:
                values_dict[parameter_name] = ['']
        print(f"{parameter_name} : {parameter_val} \n")
    return values_dict


# Function to Store Scrape Values without lossing file format
def store(values_dict, file_path = 'Book1.xlsx'):
    # Convert the dictionary to a dataframe
    scrape_data_df = pd.DataFrame(values_dict)
    scrape_data_df[['High','Low']] = scrape_data_df['High / Low'].to_list()
    scrape_data_df.drop(columns=['High / Low'], inplace=True)
    scrape_data_df

    # Load the Excel file
    wb = load_workbook(filename=file_path)

    # Select the active sheet
    sheet = wb.active

    # Update the 'Current p/e' column
    for idx in range(2, sheet.max_row + 1):

        # Assign the updated value back to the 4th colum :'Current p/e' column
        sheet.cell(row=idx, column=4).value = scrape_data_df['Stock P/E'][idx-2]
        # Assign the updated value back to the 7th colum :'Current value' column
        sheet.cell(row=idx, column=7).value = scrape_data_df['Current Price'][idx-2]
        # Assign the updated value back to the 8th colum :'High' column
        sheet.cell(row=idx, column=8).value = scrape_data_df['High'][idx-2]
        # Assign the updated value back to the 9th colum :'Low' column
        sheet.cell(row=idx, column=9).value = scrape_data_df['Low'][idx-2]
        # Assign the updated value back to the 11th colum :'Piotroski score' column
        sheet.cell(row=idx, column=11).value = scrape_data_df['Piotroski score'][idx-2]
        # Assign the updated value back to the 13th colum :'Altman Z Score' column
        sheet.cell(row=idx, column=13).value = scrape_data_df['Altman Z Score'][idx-2]
        # Assign the updated value back to the 15th colum :'G Factor' column
        sheet.cell(row=idx, column=15).value = scrape_data_df['G Factor'][idx-2]
        # Assign the updated value back to the 23th colum :'CAGR' column
        sheet.cell(row=idx, column=23).value = scrape_data_df['Return over 3years'][idx-2]
        # Assign the updated value back to the 26th colum :'Dividend Yield' column
        sheet.cell(row=idx, column=26).value = scrape_data_df['Dividend Yield'][idx-2]/100

    # Save the changes
    wb.save(file_path)


# Function to Run Scraping Process
def long_running_function():
    # load file
    company_df = pd.read_excel(file_path)
    values_dict={}

    # Initalize Chrome Driver
    driver = webdriver.Chrome(service=s, options=options)

    # Login to the website
    login(driver)

    for i,ticker in enumerate(company_df['Symbols'].to_list()):
        driver.get(f'https://www.screener.in/company/{ticker}/')
        time.sleep(2)

        # Check if the key already exists in the dictionary
        if 'Company' in values_dict:
            values_dict['Company'].append(ticker)
        else:
            values_dict['Company'] = [ticker]
        
        # get page content and extract values
        content = get_soup(driver.page_source)
        page = content.find('ul', attrs={'id':'top-ratios'})
        try:
            if len(page.find_all('li'))==13:
                values_dict = extract_values(page, values_dict)
            else:
                driver.refresh()
                time.sleep(4)
                content = get_soup(driver.page_source)
                page = content.find('ul', attrs={'id':'top-ratios'})
                values_dict = extract_values(page,values_dict)
        except:
            time.sleep(5)
            driver.refresh()
            time.sleep(3)
            content = get_soup(driver.page_source)
            page = content.find('ul', attrs={'id':'top-ratios'})
            values_dict = extract_values(page,values_dict)
        
        # Simulating a time-consuming task
        status_label.config(text=f"Running... Progress: {i+1}/{len(company_df['Symbols'].to_list())}")
        status_label.update()

    # Store Data into file
    store(values_dict, file_path)

    # Task completion message
    status_label.config(text="Task completed")
    messagebox.showinfo("Task Completed", "The task has finished successfully.")

# Create the Tkinter window
window = tk.Tk()
window.title("Automate Fundamental Data Task")
window.geometry("300x150")

# Create the button
button = tk.Button(window, text="Start Task", command=long_running_function)
button.pack(pady=20)

# Create the label
status_label = tk.Label(window, text="Click the button to start the task")
status_label.pack()

# Run the Tkinter event loop
window.mainloop()
